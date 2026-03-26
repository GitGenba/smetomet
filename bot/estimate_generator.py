import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

POSITIONS_FILE = Path(__file__).parent / "positions.json"
LOGO_FILE = Path(__file__).parent / "logo.png"


def load_positions():
    """Загружает справочник позиций из JSON"""
    with open(POSITIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def get_positions_prompt():
    """Возвращает справочник позиций в текстовом формате для промпта"""
    data = load_positions()
    lines = ["СПРАВОЧНИК УСЛУГ:\n"]
    
    current_category = None
    for service in data["services"]:
        if service["category"] != current_category:
            current_category = service["category"]
            lines.append(f"\n=== {current_category.upper()} ===")
        
        levels_info = []
        for lvl in ["1", "2", "3"]:
            if lvl in service["levels"]:
                cost = service["levels"][lvl]["cost"]
                levels_info.append(f"L{lvl}: {cost:,}₽")
        
        lines.append(f"• {service['name']} ({', '.join(levels_info)})")
    
    return "\n".join(lines)


def find_service(name: str, data: dict) -> dict | None:
    """Ищет услугу по названию (нечёткий поиск)"""
    name_lower = name.lower().strip()
    
    for service in data["services"]:
        if service["name"].lower() == name_lower:
            return service
        if name_lower in service["name"].lower():
            return service
    
    return None


def generate_estimate_excel(estimate_data: dict, output_path: str, client_mode: bool = False) -> str:
    """
    Генерирует Excel-файл сметы в стиле оригинальной таблицы Dinamika.media
    
    Args:
        estimate_data: Данные сметы
        output_path: Путь для сохранения файла
        client_mode: Если True, скрывает столбцы себестоимости/дохода и использует серые цвета
    """
    
    data = load_positions()
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    
    # ═══════════════════════════════════════════════════════════════
    # СТИЛИ (копируем из оригинала)
    # ═══════════════════════════════════════════════════════════════
    
    # Шрифты
    font_header = Font(bold=True, size=11)
    font_category = Font(bold=True, size=14)
    font_table_header = Font(bold=False, size=11)
    font_regular = Font(size=11)
    font_total = Font(bold=True, size=12)
    font_disclaimer = Font(italic=True, size=9)
    
    # Заливки (в клиентском режиме — только оттенки серого)
    fill_category = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    if client_mode:
        fill_cost_column = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        fill_total = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        fill_cost_column = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        fill_total = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматы чисел
    money_format = '#,##0" ₽"'
    percent_format = '0%'
    
    # ═══════════════════════════════════════════════════════════════
    # ШИРИНА СТОЛБЦОВ (как в оригинале)
    # ═══════════════════════════════════════════════════════════════
    
    # В клиентском режиме другая структура столбцов:
    # A - Позиция, B - Цена за ед., C - Кол-во, D - Итого
    if client_mode:
        ws.column_dimensions['A'].width = 59
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
    else:
        ws.column_dimensions['A'].width = 59
        ws.column_dimensions['B'].width = 19
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
    
    # ═══════════════════════════════════════════════════════════════
    # ЛОГОТИП
    # ═══════════════════════════════════════════════════════════════
    
    if LOGO_FILE.exists():
        img = Image(str(LOGO_FILE))
        # Оригинальный размер 311x499, масштабируем до высоты ~100px
        scale = 100 / 499
        img.width = int(311 * scale)
        img.height = int(499 * scale)
        ws.add_image(img, 'F1')
    
    # ═══════════════════════════════════════════════════════════════
    # ШАПКА ПРОЕКТА
    # ═══════════════════════════════════════════════════════════════
    
    project_name = estimate_data.get('project_name', 'Без названия')
    project_details = estimate_data.get('project_details', '')
    
    # Парсим детали проекта
    chronometry = ""
    deadline = ""
    format_info = "16:9"
    
    if project_details:
        parts = project_details.split(',')
        for part in parts:
            part = part.strip()
            if 'мин' in part.lower() or 'сек' in part.lower() or 'хрон' in part.lower():
                chronometry = part
            elif 'срок' in part.lower() or 'до ' in part.lower():
                deadline = part
            elif ':' in part and any(c.isdigit() for c in part):
                format_info = part
    
    if not chronometry:
        chronometry = project_details
    
    # Row 2: Название проекта
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Проект: {project_name}"
    ws['A2'].font = font_header
    
    # Row 3: Хронометраж
    ws.merge_cells('A3:C3')
    ws['A3'] = f"Хронометраж: {chronometry}" if chronometry else ""
    ws['A3'].font = font_header
    
    # Row 4: Срок
    ws.merge_cells('A4:C4')
    ws['A4'] = f"Срок оказания услуг: {deadline}" if deadline else ""
    ws['A4'].font = font_header
    ws['A4'].fill = fill_white
    
    # Row 5: Формат
    ws['A5'] = f"Формат: {format_info}"
    ws['A5'].fill = fill_white
    
    # Row 6: Дата сметы
    ws.merge_cells('A6:C6')
    today = datetime.now().strftime("%d.%m.%Y")
    ws['A6'] = f"Дата составления сметы: {today}"
    ws['A6'].font = font_header
    ws['A6'].fill = fill_white
    
    # Высота строк шапки
    for r in range(1, 8):
        ws.row_dimensions[r].height = 20
    
    # ═══════════════════════════════════════════════════════════════
    # ДАННЫЕ СМЕТЫ
    # ═══════════════════════════════════════════════════════════════
    
    row = 9
    current_category = None
    total_price = 0
    total_cost = 0
    total_profit = 0
    
    # Группируем items по категориям
    items_by_category = {"Препродакшн": [], "Продакшн": [], "Постпродакшн": []}
    
    for item in estimate_data.get('items', []):
        service = find_service(item['service_name'], data)
        if service:
            cat = service['category']
            if cat in items_by_category:
                items_by_category[cat].append((item, service))
    
    category_names = {
        "Препродакшн": "PREPRODUCTION",
        "Продакшн": "PRODUCTION",
        "Постпродакшн": "POSTPRODUCTION"
    }
    
    for category in ["Препродакшн", "Продакшн", "Постпродакшн"]:
        items = items_by_category[category]
        if not items:
            continue
        
        # Заголовок категории
        if client_mode:
            ws.merge_cells(f'A{row}:D{row}')
        else:
            ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value=category_names[category])
        cell.font = font_category
        cell.fill = fill_category
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Заголовки таблицы (в клиентском режиме — только 4 столбца)
        if client_mode:
            headers = ['Позиция', 'Цена за ед.', 'Кол-во', 'Итого']
        else:
            headers = ['Позиция', 'Себестоимость', 'Наценка %', 'Цена за ед.', 'Кол-во', 'Итого', 'Себест. итого', 'Доход']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = font_table_header
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Данные
        for item, service in items:
            level = str(item.get('level', 1))
            if level not in service['levels']:
                level = list(service['levels'].keys())[0]
            
            cost = service['levels'][level]['cost']
            markup = item.get('markup', 0.3)
            quantity = item.get('quantity', 1)
            unit_price = cost * (1 + markup)
            item_total = unit_price * quantity
            cost_total = cost * quantity
            profit = item_total - cost_total
            
            if client_mode:
                # Клиентский режим: только 4 столбца
                # A - Позиция
                cell = ws.cell(row=row, column=1, value=service['name'])
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # B - Цена за единицу
                cell = ws.cell(row=row, column=2, value=unit_price)
                cell.number_format = money_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                
                # C - Количество
                cell = ws.cell(row=row, column=3, value=quantity)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # D - Итого
                cell = ws.cell(row=row, column=4, value=item_total)
                cell.number_format = money_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
            else:
                # Полный режим: 8 столбцов
                # Позиция
                cell = ws.cell(row=row, column=1, value=service['name'])
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Себестоимость
                cell = ws.cell(row=row, column=2, value=cost)
                cell.number_format = money_format
                cell.border = thin_border
                cell.fill = fill_cost_column
                cell.alignment = Alignment(horizontal='right')
                
                # Наценка
                cell = ws.cell(row=row, column=3, value=markup)
                cell.number_format = percent_format
                cell.border = thin_border
                cell.fill = fill_cost_column
                cell.alignment = Alignment(horizontal='center')
                
                # Цена за единицу
                cell = ws.cell(row=row, column=4, value=unit_price)
                cell.number_format = money_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                
                # Количество
                cell = ws.cell(row=row, column=5, value=quantity)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Итого
                cell = ws.cell(row=row, column=6, value=item_total)
                cell.number_format = money_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                
                # Себестоимость итого
                cell = ws.cell(row=row, column=7, value=cost_total)
                cell.number_format = money_format
                cell.border = thin_border
                cell.fill = fill_cost_column
                cell.alignment = Alignment(horizontal='right')
                
                # Доход
                cell = ws.cell(row=row, column=8, value=profit)
                cell.number_format = money_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
            
            total_price += item_total
            total_cost += cost_total
            total_profit += profit
            row += 1
        
        row += 1  # Пустая строка между категориями
    
    # ═══════════════════════════════════════════════════════════════
    # ИТОГО
    # ═══════════════════════════════════════════════════════════════
    
    row += 1
    
    if client_mode:
        # Клиентский режим: ИТОГО в 4 столбцах
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws.cell(row=row, column=1, value="ИТОГО:")
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        cell = ws.cell(row=row, column=4, value=total_price)
        cell.number_format = money_format
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
    else:
        # Полный режим: ИТОГО в 8 столбцах
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value="ИТОГО:")
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        cell = ws.cell(row=row, column=6, value=total_price)
        cell.number_format = money_format
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
        
        cell = ws.cell(row=row, column=7, value=total_cost)
        cell.number_format = money_format
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
        
        cell = ws.cell(row=row, column=8, value=total_profit)
        cell.number_format = money_format
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
    
    ws.row_dimensions[row].height = 25
    
    # ═══════════════════════════════════════════════════════════════
    # ДИСКЛЕЙМЕР
    # ═══════════════════════════════════════════════════════════════
    
    row += 3
    if client_mode:
        ws.merge_cells(f'A{row}:D{row}')
    else:
        ws.merge_cells(f'A{row}:H{row}')
    disclaimer = "ВАЖНО: Оценка стоимости проекта проводилась на основе полученного брифа. В смете указаны все позиции, необходимые для реализации проекта и сдачи финального видеоматериала. В случае изменения технического задания, менеджер укажет Вам на возможность добавления/исключения тех или иных позиций. С Уважением, команда Dinamika.media"
    cell = ws.cell(row=row, column=1, value=disclaimer)
    cell.font = font_disclaimer
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 45
    
    wb.save(output_path)
    return output_path


def format_estimate_text(estimate_data: dict) -> str:
    """Форматирует смету в текстовый вид для Telegram"""
    data = load_positions()
    
    lines = [
        f"📋 *{estimate_data.get('project_name', 'Смета проекта')}*",
        f"_{estimate_data.get('project_details', '')}_",
        "",
        "━━━━━━━━━━━━━━━━━━━━━━"
    ]
    
    current_category = None
    total_price = 0
    
    for item in estimate_data.get('items', []):
        service = find_service(item['service_name'], data)
        if not service:
            continue
        
        if service['category'] != current_category:
            current_category = service['category']
            lines.append(f"\n*{current_category.upper()}*")
        
        level = str(item.get('level', 1))
        if level not in service['levels']:
            level = list(service['levels'].keys())[0]
        
        cost = service['levels'][level]['cost']
        markup = item.get('markup', 0.3)
        quantity = item.get('quantity', 1)
        unit_price = cost * (1 + markup)
        total = unit_price * quantity
        
        qty_str = f" × {quantity}" if quantity > 1 else ""
        lines.append(f"• {service['name']}{qty_str}: *{total:,.0f}₽*")
        total_price += total
    
    lines.extend([
        "",
        "━━━━━━━━━━━━━━━━━━━━━━",
        f"💰 *ИТОГО: {total_price:,.0f}₽*"
    ])
    
    return "\n".join(lines)
